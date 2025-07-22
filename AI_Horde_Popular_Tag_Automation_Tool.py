import os
import requests
import json
import pandas as pd
import re

# Constants
API_URL = "https://aihorde.net/api/v2/stats/text/models"

# Directory where input and output files are stored
USER_FILES_DIR = os.path.join(os.path.dirname(__file__), "user-files")
os.makedirs(USER_FILES_DIR, exist_ok=True)

USAGE_JSON = os.path.join(USER_FILES_DIR, "RawModelUsageData.json")
MODELS_CSV = os.path.join(USER_FILES_DIR, "models.csv")
OUTPUT_CSV = os.path.join(USER_FILES_DIR, "models_updated.csv")
TOP_N = 25
PERIOD_TAGS = {
    'day': 'daily_top25',
    'month': 'monthly_top25',
    'total': 'alltime_top25'
}

# Step 1: Fetch latest usage data
print(f"Fetching usage data from {API_URL}...")
r = requests.get(API_URL)
r.raise_for_status()
usage_data = r.json()

# Optionally save raw JSON
with open(USAGE_JSON, 'w', encoding='utf-8') as f:
    json.dump(usage_data, f, indent=2)

# Step 2: Build DataFrame of usage
records = []
for period, models in usage_data.items():
    for model_name, count in models.items():
        records.append({'period': period, 'model': model_name, 'usage_count': count})

df_usage = pd.DataFrame(records)
df_usage.to_csv(os.path.join(USER_FILES_DIR, "usage_data.csv"), index=False)

# Convert the usage data to an Excel workbook with one sheet per period
usage_xlsx = os.path.join(USER_FILES_DIR, "usage_data.xlsx")
with pd.ExcelWriter(usage_xlsx) as writer:
    for period, group in df_usage.groupby("period"):
        sheet = period.capitalize()
        # Include headers when writing each sheet
        group.drop(columns="period").to_excel(writer, index=False, sheet_name=sheet)

# Format each sheet as a table once written
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

wb = load_workbook(usage_xlsx)
for sheet in wb.sheetnames:
    ws = wb[sheet]
    max_row = ws.max_row
    max_col = ws.max_column
    end_col = get_column_letter(max_col)
    table_ref = f"A1:{end_col}{max_row}"
    table = Table(displayName=f"{sheet}Table", ref=table_ref)
    style = TableStyleInfo(name="TableStyleMedium9", showFirstColumn=False,
                           showLastColumn=False, showRowStripes=True, showColumnStripes=False)
    table.tableStyleInfo = style
    ws.add_table(table)

# Auto-fit column widths for each sheet
for sheet in wb.sheetnames:
    ws = wb[sheet]
    for col in ws.columns:
        max_length = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            try:
                cell_length = len(str(cell.value)) if cell.value is not None else 0
                if cell_length > max_length:
                    max_length = cell_length
            except Exception:
                pass
        adjusted_width = max_length + 2  # Add some padding
        ws.column_dimensions[col_letter].width = adjusted_width

wb.save(usage_xlsx)

# Step 3: Clean model names and merge duplicates
print("Cleaning model names and merging duplicates...")

# Load whitelist for proper capitalization if available
if os.path.exists(MODELS_CSV):
    models_df = pd.read_csv(MODELS_CSV)
    models_df['model_short'] = models_df['name'].astype(str).str.split('/').str[-1]
    model_short_map = {ms.lower(): ms for ms in models_df['model_short']}
else:
    model_short_map = {}

# Regex used elsewhere in the Horde tooling to strip quantization suffixes
# from model names (e.g. "-Q4_K_M", "-QInt8", etc.)
official_quant_regex = re.compile(r'-[a-zA-Z]+?Q(-[Ii]nt)?[0-9]([_-][a-zA-Z]+)*$', re.IGNORECASE)

# Additional pattern to strip other trailing tokens such as ".i1" or "-exl2"
extra_suffix_regex = re.compile(r'(?i)([._-](?:iMat|iMatrix|i\d+|b\d+|c\d+|ch\d+|bpw|h\d+|exl\d+).*)$')

def strip_quant(name: str) -> str:
    # remove quantization descriptors
    name = official_quant_regex.sub('', name)
    # remove any remaining suffixes like .i1 or -exl2
    return extra_suffix_regex.sub('', name)

def map_to_whitelist(raw_short: str):
    raw_lower = raw_short.lower()
    for short_lower, short_orig in model_short_map.items():
        if raw_lower.endswith(short_lower):
            return short_orig
    return None

cleaned_dfs = {}
for sheet in wb.sheetnames:
    df = pd.read_excel(usage_xlsx, sheet_name=sheet)
    df['raw_short'] = df['model'].astype(str).str.split('/').str[-1]
    df['mapped'] = df['raw_short'].apply(map_to_whitelist)
    df['whitelisted'] = df['mapped'].notnull()
    df['cleaned'] = df['raw_short'].apply(strip_quant)
    df['model'] = df.apply(lambda r: r['mapped'] if pd.notnull(r['mapped']) else r['cleaned'], axis=1)
    df_merged = df.groupby('model', as_index=False).agg({'usage_count': 'sum', 'whitelisted': 'max'})
    df_merged['whitelisted'] = df_merged['whitelisted'].apply(lambda x: 'T' if x else 'F')
    cleaned_dfs[sheet] = df_merged

with pd.ExcelWriter(usage_xlsx, engine='openpyxl') as writer:
    for sheet, cdf in cleaned_dfs.items():
        cdf.to_excel(writer, index=False, sheet_name=sheet)

wb = load_workbook(usage_xlsx)
for sheet in wb.sheetnames:
    ws = wb[sheet]
    max_row = ws.max_row
    max_col = ws.max_column
    end_col = get_column_letter(max_col)
    table_ref = f"A1:{end_col}{max_row}"
    table = Table(displayName=f"{sheet}Table", ref=table_ref)
    style = TableStyleInfo(name="TableStyleMedium9", showFirstColumn=False,
                           showLastColumn=False, showRowStripes=True, showColumnStripes=False)
    table.tableStyleInfo = style
    ws.add_table(table)

for sheet in wb.sheetnames:
    ws = wb[sheet]
    for col in ws.columns:
        max_length = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            try:
                cell_length = len(str(cell.value)) if cell.value is not None else 0
                if cell_length > max_length:
                    max_length = cell_length
            except Exception:
                pass
        ws.column_dimensions[col_letter].width = max_length + 2

wb.save(usage_xlsx)

## STOP HERE FOR NOW ##
## Below this are extra features that are broken, and I don't want them anyway ##
# # Step 3: Determine top N models per period
# top_models = {}
# for period, group in df_usage.groupby('period'):
#     # Sort by usage and take top N
#     top_list = group.sort_values('usage_count', ascending=False)
#     top_models[period] = set(top_list.head(TOP_N)['model'].tolist())
#     print(f"Top {TOP_N} for {period}: {len(top_models[period])} models")


# # Step 4: Load existing models whitelist
# df_models = pd.read_csv(MODELS_CSV)
# # Ensure tags column exists
# if 'tags' not in df_models.columns:
#     df_models['tags'] = ''

# # Step 5: Update tags based on membership in top lists
# new_tags = []
# for _, row in df_models.iterrows():
#     model = row['name']
#     tags = []
#     for period, tag in PERIOD_TAGS.items():
#         if model in top_models.get(period, set()):
#             tags.append(tag)
#     # Join with commas, preserve any existing tags if desired
#     new_tags.append(','.join(tags))

# df_models['tags'] = new_tags

# # Step 6: Save updated CSV
# df_models.to_csv(OUTPUT_CSV, index=False)
# print(f"Updated CSV written to {OUTPUT_CSV}")
